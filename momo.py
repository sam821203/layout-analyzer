# 抓 momo 活動頁面
# 目標 1：抓到頁面使用的宮格
# 目標 2：將抓到的宮格轉成一張粗略的 wireframe

import urllib.request as req
import bs4
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 抓取 momo 活動頁網頁原始碼
url = 'https://www.momoshop.com.tw/edm/cmmedm.jsp?lpn=O3XZkrInHiH&n=1'

# 建立 request 物件，附加 request headers 的資訊
request = req.Request(url, headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36"
})
with req.urlopen(request) as response:
    data = response.read().decode('utf-8')
# print(data)

# 解析網頁 HTML 結構
root = bs4.BeautifulSoup(data, 'html.parser')

# 找到活動頁名稱
spName = root.title.string
# allPDLayout = root.find_all(
#     'div', 
#     class_='PD_layout', 
#     attrs={'class': lambda e: e.startswith('layout_')}
# )

allPDLayout = root.select('div[class^="PD_layout layout_"]')
colDivs = []

for pdLayout in allPDLayout:
    # 判斷屬性值是否為空值或非數字
    # 獲取屬性值
    hasAttribute = pdLayout.has_attr('data-pd-col-pc')
    deskColNum = pdLayout.get('data-pd-col-pc')
    
    if not hasAttribute or not deskColNum.isdigit():
        print('data-pd-col-pc 的值有誤')
    else:
        # 將抓到的 deskColNum 塞進去 template，並塞進去 html
        template = f'<div data-col="{deskColNum}"></div>'
        colDivs.append(template)
        # print(f'商品排站 {deskColNum}x')

# 使用 join() 方法將列表中的字串合併為一個完整的 HTML 字串
html = '\n'.join(colDivs)

with open('index.html', 'w') as f:
    f.write(html)

# # 創建一個 Excel 檔案
# wb = Workbook()

# # 選擇要編輯的工作表
# ws = wb.active

# # 創建一個填充樣式，例如藍色填充
# tomatoFill = PatternFill(
#     start_color='F8CBAD', 
#     end_color='F8CBAD', 
#     fill_type='solid'
# )

# # 將資訊寫入 Excel 檔案
# cell = ws['A1']
# cell.value = '活動頁名稱'
# cell.fill = tomatoFill
# cell.alignment = Alignment(horizontal = 'center', vertical = 'center')

# # 調整 cell 儲存格的寬度以適應文字長度
# cellWidth = len(cell.value) * 3
# ws.column_dimensions['A'].width = cellWidth

# ws['A2'] = spName

# # 儲存 Excel 檔案
# wb.save('momo.xlsx')